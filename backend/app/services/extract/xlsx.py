from __future__ import annotations

from ._common import rows_to_table


def extract_xlsx(file_path: str, xls: bool = False) -> dict:
    if xls:
        return _extract_xls(file_path)
    return _extract_xlsx(file_path)


def _extract_xlsx(file_path: str) -> dict:
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames
        tables: list[dict] = []
        text_parts: list[str] = []
        total_rows = 0
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = list(row)
                if any(value is not None and str(value).strip() for value in values):
                    rows.append(values)
            if not rows:
                continue
            table = rows_to_table(sheet_name, rows)
            tables.append(table)
            total_rows += len(table["rows"])
            header = table["headers"] or rows[0]
            text_parts.append(f"Sheet: {sheet_name}")
            if header:
                text_parts.append(", ".join(value or "" for value in header))
    finally:
        workbook.close()

    return {
        "text": "\n".join(text_parts),
        "tables": tables[:_MAX_TABLES],
        "metadata": {"sheetCount": len(sheet_names), "rowCount": total_rows},
    }


def _extract_xls(file_path: str) -> dict:
    import xlrd

    workbook = xlrd.open_workbook(file_path)
    tables: list[dict] = []
    text_parts: list[str] = []
    total_rows = 0
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            values = [
                sheet.cell_value(row_index, col_index)
                for col_index in range(sheet.ncols)
            ]
            if any(str(value).strip() for value in values):
                rows.append(values)
        if not rows:
            continue
        table = rows_to_table(sheet.name, rows)
        tables.append(table)
        total_rows += len(table["rows"])
        header = table["headers"] or rows[0]
        text_parts.append(f"Sheet: {sheet.name}")
        if header:
            text_parts.append(", ".join(str(value) for value in header))

    return {
        "text": "\n".join(text_parts),
        "tables": tables[:_MAX_TABLES],
        "metadata": {
            "sheetCount": len(workbook.sheets()),
            "rowCount": total_rows,
        },
    }


_MAX_TABLES = 50
